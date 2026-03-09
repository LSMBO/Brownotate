import openpyxl
from openpyxl.styles import Font
from openpyxl.styles import PatternFill
from openpyxl.styles import Border, Side
import utils

taxid2name = utils.get_taxid_to_scientificname()

def write_excel(data, filename, header_bg="eeffed"):
    lengths = [len(v) for v in data.values()]
    if len(set(lengths)) != 1:
        length_check = {k: len(v) for k, v in data.items()}
        print(f"Column lengths: {length_check}")
        raise ValueError("All columns must have the same number of rows.")
    
    wb = openpyxl.Workbook()
    ws = wb.active
    headers = list(data.keys())
    ws.append(headers)
    for row in range(len(next(iter(data.values())))):
        ws.append([data[header][row] for header in headers])
    border = Side(border_style="thin")
    for cell in ws["1:1"]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill('solid', start_color=header_bg)
        cell.border = Border(left=border, right=border, top=border, bottom=border)
        
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            if cell.value and len(str(cell.value)) > max_length:
                max_length = len(cell.value)
        ws.column_dimensions[column_letter].width = min(50, max_length + 2)    
    wb.save(filename)

def add_sheet(data, filename, sheet_name, header_bg="eeffed"):
    lengths = [len(v) for v in data.values()]
    if len(set(lengths)) != 1:
        length_check = {k: len(v) for k, v in data.items()}
        print(f"Column lengths: {length_check}")
        raise ValueError("All columns must have the same number of rows.")
    
    try:
        wb = openpyxl.load_workbook(filename)
    except FileNotFoundError:
        wb = openpyxl.Workbook()
        if len(wb.sheetnames) == 1 and wb.active.max_row <= 1:
            wb.remove(wb.active)
    
    if sheet_name in wb.sheetnames:
        del wb[sheet_name]
    
    ws = wb.create_sheet(sheet_name)
    
    headers = list(data.keys())
    ws.append(headers)
    for row in range(len(next(iter(data.values())))):
        ws.append([data[header][row] for header in headers])
    
    border = Side(border_style="thin")
    for cell in ws["1:1"]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill('solid', start_color=header_bg)
        cell.border = Border(left=border, right=border, top=border, bottom=border)
    
    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            if cell.value and len(str(cell.value)) > max_length:
                max_length = len(str(cell.value))
        ws.column_dimensions[column_letter].width = min(50, max_length + 2)
    
    wb.save(filename)


def add_hit(output_data, hit):
    output_data["Query accession"].append(hit.get("qseqid", ""))
    output_data["Subject accession"].append(hit.get("sseqid",""))
    output_data["Subject description"].append(hit.get("stitle",""))
    output_data["Subject species (taxid)"].append(str(hit.get("staxid") or ""))
    output_data["Subject species (name)"].append(taxid2name.get(str(hit.get("staxid")), "") if hit.get("staxid") else "")
    output_data["Gene Name"].append(utils.gene_name_from_stitle(hit.get("stitle","")))
    output_data["Bitscore"].append(f"{hit.get('bits',0):.1f}")
    output_data["Evalue"].append(f"{hit.get('evalue',0):.1e}")
    output_data["Identity (%)"].append(f"{hit.get('pident',0):.2f}")
    output_data["Similarity (%)"].append(f"{hit.get('ppos',0):.2f}") 
    output_data["Query coverage (%)"].append(f"{hit.get('_qcov',0)*100:.2f}")
    output_data["Subject coverage (%)"].append(f"{hit.get('_scov',0)*100:.2f}")
    output_data["Common ancestor (rank)"].append(hit.get("common_ancestor_rank",""))
    output_data["Common ancestor (taxID)"].append(hit.get("common_ancestor_taxid",""))
    output_data["Common ancestor (name)"].append(hit.get("common_ancestor_name",""))
    output_data["Hit found"].append("True")    

    return output_data


def add_no_hit(output_data, qid):
    output_data["Query accession"].append(qid)
    output_data["Subject accession"].append("")
    output_data["Subject description"].append("")
    output_data["Subject species (taxid)"].append("")
    output_data["Subject species (name)"].append("")
    output_data["Gene Name"].append("")
    output_data["Bitscore"].append("")
    output_data["Evalue"].append("")
    output_data["Identity (%)"].append("")
    output_data["Similarity (%)"].append("")
    output_data["Query coverage (%)"].append("")
    output_data["Subject coverage (%)"].append("")
    output_data["Common ancestor (rank)"].append("")
    output_data["Common ancestor (taxID)"].append("")
    output_data["Common ancestor (name)"].append("")
    output_data["Hit found"].append("False")
    
    return output_data


